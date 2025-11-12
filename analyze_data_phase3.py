#!/usr/bin/env python3
"""
Phase 3: Export complete data for dashboard
"""

import json
import openpyxl
from datetime import datetime
from collections import Counter

file_path = '/vercel/sandbox/uploads/عقود الصيانة والتركيب حتى تاريخ 15-10-2025.xlsx'
print(f"\n{'='*60}")
print("PHASE 3: EXPORTING COMPLETE DATA")
print(f"{'='*60}\n")

wb = openpyxl.load_workbook(file_path, data_only=True)

def clean_value(val):
    """Clean and normalize values"""
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    return str(val).strip()

def export_sheet(ws, sheet_name):
    """Export complete sheet data"""
    print(f"Exporting '{sheet_name}'...")
    
    # Get headers
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        header = str(cell).strip() if cell else ""
        if header and not header.startswith('Column_'):
            headers.append(header)
    
    # Collect all data
    all_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell for cell in row):
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = clean_value(cell)
            all_rows.append(row_dict)
    
    print(f"  Exported {len(all_rows)} rows")
    return all_rows

# Export all sheets
complete_data = {
    'export_timestamp': datetime.now().isoformat(),
    'file_name': 'عقود الصيانة والتركيب حتى تاريخ 15-10-2025.xlsx',
    'sheets': {}
}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    complete_data['sheets'][sheet_name] = export_sheet(ws, sheet_name)

# Calculate comprehensive statistics for maintenance sheet (main data)
maintenance_data = complete_data['sheets']['maintenance']

print(f"\n{'='*60}")
print("CALCULATING STATISTICS")
print(f"{'='*60}\n")

stats = {
    'total_contracts': len(maintenance_data),
    'regions': {},
    'prices': {},
    'dates': {},
    'equipment': {},
    'employees': {}
}

# Region analysis
regions = [row.get('REGN_NM', '') for row in maintenance_data if row.get('REGN_NM')]
region_counts = Counter(regions)
stats['regions'] = {
    'total_unique': len(region_counts),
    'distribution': dict(region_counts.most_common(15)),
    'top_region': region_counts.most_common(1)[0] if region_counts else ('N/A', 0)
}

# Price analysis
prices = []
for row in maintenance_data:
    try:
        price_str = str(row.get('Prc', '0')).replace(',', '').strip()
        if price_str:
            price = float(price_str)
            if price > 0:
                prices.append(price)
    except Exception as e:
        pass

if prices:
    stats['prices'] = {
        'total_revenue': sum(prices),
        'average_price': sum(prices) / len(prices),
        'min_price': min(prices),
        'max_price': max(prices),
        'count': len(prices)
    }
else:
    stats['prices'] = {
        'total_revenue': 0,
        'average_price': 0,
        'min_price': 0,
        'max_price': 0,
        'count': 0
    }

# Stops analysis
stops = [row.get('stops', '') for row in maintenance_data if row.get('stops')]
stops_counts = Counter(stops)
stats['equipment']['stops_distribution'] = dict(stops_counts.most_common(10))

# Control type analysis
control_types = [row.get('Cntrl_ty', '') for row in maintenance_data if row.get('Cntrl_ty')]
control_counts = Counter(control_types)
stats['equipment']['control_types'] = dict(control_counts.most_common(10))

# Model analysis
models = [row.get('Mdl', '') for row in maintenance_data if row.get('Mdl')]
model_counts = Counter(models)
stats['equipment']['models'] = dict(model_counts.most_common(10))

# Employee analysis
employees = [row.get('emp_nm_ar', '') for row in maintenance_data if row.get('emp_nm_ar')]
employee_counts = Counter(employees)
stats['employees'] = {
    'total_unique': len(employee_counts),
    'distribution': dict(employee_counts.most_common())
}

# Date range analysis
start_dates = [row.get('CurStrt_dt', '') for row in maintenance_data if row.get('CurStrt_dt')]
end_dates = [row.get('curend_dt', '') for row in maintenance_data if row.get('curend_dt')]

if start_dates:
    stats['dates']['earliest_start'] = min(start_dates)
    stats['dates']['latest_start'] = max(start_dates)
if end_dates:
    stats['dates']['earliest_end'] = min(end_dates)
    stats['dates']['latest_end'] = max(end_dates)

# Contract status (active/expired based on end date)
from datetime import datetime as dt
today = dt.now().strftime('%Y-%m-%d')
active_contracts = sum(1 for row in maintenance_data if row.get('curend_dt', '') >= today)
expired_contracts = len(maintenance_data) - active_contracts

stats['contract_status'] = {
    'active': active_contracts,
    'expired': expired_contracts,
    'total': len(maintenance_data)
}

complete_data['statistics'] = stats

# Save complete data
with open('/vercel/sandbox/public/contracts_data.json', 'w', encoding='utf-8') as f:
    json.dump(complete_data, f, ensure_ascii=False, indent=2)

print(f"\n{'='*60}")
print("PHASE 3 COMPLETE")
print(f"{'='*60}\n")
print("✓ Complete data exported to public/contracts_data.json")
print(f"\nKey Statistics:")
print(f"  Total Contracts: {stats['total_contracts']}")
print(f"  Total Revenue: {stats['prices']['total_revenue']:,.0f} KWD")
print(f"  Average Price: {stats['prices']['average_price']:,.0f} KWD")
print(f"  Unique Regions: {stats['regions']['total_unique']}")
print(f"  Active Contracts: {stats['contract_status']['active']}")
print(f"  Expired Contracts: {stats['contract_status']['expired']}")
print(f"  Top Region: {stats['regions']['top_region'][0]} ({stats['regions']['top_region'][1]} contracts)")
