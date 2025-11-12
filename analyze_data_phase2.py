#!/usr/bin/env python3
"""
Phase 2: Deep analysis - unique values, stats, patterns
"""

import json
import openpyxl
from collections import Counter, defaultdict
from datetime import datetime
import re

file_path = '/vercel/sandbox/uploads/عقود الصيانة والتركيب حتى تاريخ 15-10-2025.xlsx'
print(f"\n{'='*60}")
print("PHASE 2: DEEP ANALYSIS")
print(f"{'='*60}\n")

wb = openpyxl.load_workbook(file_path, data_only=True)

def clean_value(val):
    """Clean and normalize values"""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    return str(val).strip()

def analyze_sheet(ws, sheet_name):
    """Perform deep analysis on a sheet"""
    print(f"\n--- Deep Analysis: '{sheet_name}' ---")
    
    # Get headers
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        headers.append(str(cell).strip() if cell else "")
    
    # Collect all data
    all_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell for cell in row):
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = clean_value(cell)
            all_rows.append(row_dict)
    
    print(f"  Total records: {len(all_rows)}")
    
    # Analyze each column
    column_analysis = {}
    
    for header in headers:
        if not header or header.startswith('Column_'):
            continue
            
        values = [row.get(header) for row in all_rows if row.get(header)]
        unique_values = set(values)
        
        analysis = {
            'total_count': len(all_rows),
            'non_empty_count': len(values),
            'empty_count': len(all_rows) - len(values),
            'unique_count': len(unique_values),
            'sample_values': list(unique_values)[:10]
        }
        
        # Try to detect numeric columns
        numeric_values = []
        for v in values:
            try:
                # Remove any non-numeric characters except decimal point
                clean_num = re.sub(r'[^\d.]', '', str(v))
                if clean_num:
                    numeric_values.append(float(clean_num))
            except:
                pass
        
        if numeric_values and len(numeric_values) > len(values) * 0.5:
            analysis['is_numeric'] = True
            analysis['min'] = min(numeric_values)
            analysis['max'] = max(numeric_values)
            analysis['avg'] = sum(numeric_values) / len(numeric_values)
            analysis['sum'] = sum(numeric_values)
        else:
            analysis['is_numeric'] = False
        
        # Get top values
        if len(unique_values) < 100:
            value_counts = Counter(values)
            analysis['top_values'] = dict(value_counts.most_common(10))
        
        column_analysis[header] = analysis
        
        print(f"  {header}: {len(values)} non-empty, {len(unique_values)} unique")
    
    return {
        'total_rows': len(all_rows),
        'headers': headers,
        'column_analysis': column_analysis,
        'sample_data': all_rows[:5]
    }

# Analyze all sheets
results = {
    'analysis_phase': 2,
    'timestamp': datetime.now().isoformat(),
    'sheets': {}
}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    results['sheets'][sheet_name] = analyze_sheet(ws, sheet_name)

# Save results
with open('/vercel/sandbox/data_analysis_phase2.json', 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

print(f"\n{'='*60}")
print("PHASE 2 COMPLETE - Deep analysis done")
print(f"{'='*60}\n")
print("✓ Results saved to data_analysis_phase2.json")
print("\nKey findings:")
for sheet_name, data in results['sheets'].items():
    print(f"  {sheet_name}: {data['total_rows']} rows, {len(data['headers'])} columns")
