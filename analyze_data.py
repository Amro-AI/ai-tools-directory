#!/usr/bin/env python3
"""
Comprehensive data analysis for maintenance and installation contracts
Phase 1: Load complete file and discover structure
"""

import json
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl

# Load the complete Excel file
file_path = '/vercel/sandbox/uploads/عقود الصيانة والتركيب حتى تاريخ 15-10-2025.xlsx'
print(f"\n{'='*60}")
print("PHASE 1: LOADING AND STRUCTURE DISCOVERY")
print(f"{'='*60}\n")

wb = openpyxl.load_workbook(file_path, data_only=True)
print(f"✓ Loaded workbook with {len(wb.sheetnames)} sheets: {wb.sheetnames}")

# Analyze each sheet
all_data = {}
sheet_info = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Analyzing Sheet: '{sheet_name}' ---")
    
    # Get all data
    data = []
    headers = []
    
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            # First row is headers
            headers = [str(cell).strip() if cell else f"Column_{i}" for i, cell in enumerate(row)]
            print(f"  Headers ({len(headers)}): {headers[:5]}...")
        else:
            # Data rows
            if any(cell for cell in row):  # Skip empty rows
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        # Clean up the data
                        if cell is not None:
                            if isinstance(cell, datetime):
                                row_dict[headers[i]] = cell.strftime('%Y-%m-%d')
                            else:
                                row_dict[headers[i]] = str(cell).strip()
                        else:
                            row_dict[headers[i]] = ""
                data.append(row_dict)
    
    all_data[sheet_name] = {
        'headers': headers,
        'data': data,
        'row_count': len(data)
    }
    
    print(f"  Total rows: {len(data)}")
    print(f"  Sample row keys: {list(data[0].keys())[:5] if data else 'No data'}")

# Save initial structure
output = {
    'analysis_phase': 1,
    'timestamp': datetime.now().isoformat(),
    'file_info': {
        'sheets': list(all_data.keys()),
        'total_sheets': len(all_data)
    },
    'sheet_details': {}
}

for sheet_name, sheet_data in all_data.items():
    output['sheet_details'][sheet_name] = {
        'row_count': sheet_data['row_count'],
        'column_count': len(sheet_data['headers']),
        'headers': sheet_data['headers']
    }

print(f"\n{'='*60}")
print("PHASE 1 COMPLETE - Structure discovered")
print(f"{'='*60}\n")

# Save to JSON for dashboard
with open('/vercel/sandbox/data_analysis.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print("✓ Phase 1 results saved to data_analysis.json")
print("\nNext: Run Phase 2 for deep analysis...")
